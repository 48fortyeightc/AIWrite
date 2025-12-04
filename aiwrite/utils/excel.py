"""
Excel 表格解析工具

支持读取 .xls 和 .xlsx 文件，转换为表格数据
"""

from __future__ import annotations

from pathlib import Path
from typing import Any


def read_excel_file(file_path: str | Path) -> list[list[str]]:
    """
    读取 Excel 文件内容
    
    Args:
        file_path: Excel 文件路径
        
    Returns:
        二维列表，每个内层列表代表一行
    """
    file_path = Path(file_path)
    
    if not file_path.exists():
        raise FileNotFoundError(f"Excel 文件不存在: {file_path}")
    
    suffix = file_path.suffix.lower()
    
    if suffix == ".xlsx":
        return _read_xlsx(file_path)
    elif suffix == ".xls":
        return _read_xls(file_path)
    else:
        raise ValueError(f"不支持的文件格式: {suffix}")


def _read_xlsx(file_path: Path) -> list[list[str]]:
    """读取 .xlsx 文件"""
    from openpyxl import load_workbook
    
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    
    rows: list[list[str]] = []
    for row in ws.iter_rows():
        row_data = []
        for cell in row:
            value = cell.value
            if value is None:
                row_data.append("")
            else:
                row_data.append(str(value))
        # 跳过完全空白的行
        if any(cell.strip() for cell in row_data):
            rows.append(row_data)
    
    wb.close()
    return rows


def _read_xls(file_path: Path) -> list[list[str]]:
    """读取 .xls 文件"""
    import xlrd
    
    wb = xlrd.open_workbook(file_path)
    ws = wb.sheet_by_index(0)
    
    rows: list[list[str]] = []
    for row_idx in range(ws.nrows):
        row_data = []
        for col_idx in range(ws.ncols):
            value = ws.cell_value(row_idx, col_idx)
            if value is None:
                row_data.append("")
            else:
                row_data.append(str(value))
        # 跳过完全空白的行
        if any(cell.strip() for cell in row_data):
            rows.append(row_data)
    
    return rows


def table_to_markdown(rows: list[list[str]]) -> str:
    """
    将表格数据转换为 Markdown 格式
    
    Args:
        rows: 二维列表，第一行为表头
        
    Returns:
        Markdown 格式的表格字符串
    """
    if not rows:
        return ""
    
    # 第一行是表头
    header = rows[0]
    
    # 计算每列的最大宽度
    col_widths = [len(h) for h in header]
    for row in rows[1:]:
        for i, cell in enumerate(row):
            if i < len(col_widths):
                col_widths[i] = max(col_widths[i], len(cell))
    
    # 构建 Markdown 表格
    lines = []
    
    # 表头
    header_line = "| " + " | ".join(h.ljust(w) for h, w in zip(header, col_widths)) + " |"
    lines.append(header_line)
    
    # 分隔线
    separator = "| " + " | ".join("-" * w for w in col_widths) + " |"
    lines.append(separator)
    
    # 数据行
    for row in rows[1:]:
        # 确保行长度与表头一致
        padded_row = row + [""] * (len(header) - len(row))
        row_line = "| " + " | ".join(c.ljust(w) for c, w in zip(padded_row, col_widths)) + " |"
        lines.append(row_line)
    
    return "\n".join(lines)


def table_to_latex(rows: list[list[str]]) -> str:
    """
    将表格数据转换为 LaTeX 格式
    
    Args:
        rows: 二维列表，第一行为表头
        
    Returns:
        LaTeX 格式的表格字符串
    """
    if not rows:
        return ""
    
    num_cols = len(rows[0])
    col_spec = "|" + "c|" * num_cols
    
    lines = [
        "\\begin{table}[htbp]",
        "\\centering",
        f"\\begin{{tabular}}{{{col_spec}}}",
        "\\hline",
    ]
    
    # 表头（加粗）
    header = rows[0]
    header_line = " & ".join(f"\\textbf{{{_escape_latex(h)}}}" for h in header) + " \\\\"
    lines.append(header_line)
    lines.append("\\hline")
    
    # 数据行
    for row in rows[1:]:
        # 确保行长度与表头一致
        padded_row = row + [""] * (num_cols - len(row))
        row_line = " & ".join(_escape_latex(c) for c in padded_row) + " \\\\"
        lines.append(row_line)
    
    lines.extend([
        "\\hline",
        "\\end{tabular}",
        "\\end{table}",
    ])
    
    return "\n".join(lines)


def _escape_latex(text: str) -> str:
    """转义 LaTeX 特殊字符"""
    replacements = [
        ("\\", "\\textbackslash{}"),
        ("&", "\\&"),
        ("%", "\\%"),
        ("$", "\\$"),
        ("#", "\\#"),
        ("_", "\\_"),
        ("{", "\\{"),
        ("}", "\\}"),
        ("~", "\\textasciitilde{}"),
        ("^", "\\textasciicircum{}"),
    ]
    for old, new in replacements:
        text = text.replace(old, new)
    return text
